# This program uses the Student class to import data from the Excel spreadsheet
# and save the data as a list of Student objects which can then be used
# for further analysis of the data.  Also, divides data into test and training
# sets, and saved data sets to file.

import openpyxl #used for accessing spreadsheet
import pickle #used for saving data to file
import nltk #used for tokenizing student essays
from student_class import Student #student class definition
from scipy import stats #used for calculating percentiles

# load data spreadsheet
wb = openpyxl.load_workbook('Learners_ALL_FINAL.xlsx')

# open the correct data table in spreadsheet
sheet = wb.get_sheet_by_name('Good Data')
# initialize empty lists to hold data sets
student_train_list = list()
student_test_list = list()
student_list= list()

# iterate through the rows in the data table.  Create a new Student object for
# each row.  Set the Student object's attributes according to the values
# in the corresponding column for that row.  In addition, calculate tokens and
# types and update the values of those attributes.
row_count = sheet.max_row
for i in range(2, row_count+1):
    student = Student()
    student.sex = sheet.cell(row=i, column=4).value
    if sheet.cell(row=i, column=6).value != ' ' and sheet.cell(row=i, column=6).value != None:
        student.age = int(sheet.cell(row=i, column=6).value)
    student.degree = sheet.cell(row=i, column=7).value
    student.native_language = sheet.cell(row=i, column=9).value
    if sheet.cell(row=i, column=11).value == 'Yes':
        student.stay_spanish_country = True
    else: student.stay_spanish_country = False
    if sheet.cell(row=i, column=19).value != ' ' and sheet.cell(row=i, column=19).value != None:
        student.age_started_spanish = float(sheet.cell(row=i, column=19).value)
    if sheet.cell(row=i, column=20).value != ' 'and sheet.cell(row=i, column=20).value != None:
        student.years_studying_spanish = float(sheet.cell(row=i, column=20).value)
    if sheet.cell(row=i, column=25).value == 'Yes':
        student.other_languages = True
    else: student.other_languages = False
    student.placement_raw = int(sheet.cell(row=i, column=36).value)
    student.placement_percent = float(sheet.cell(row=i, column=37).value)
    student.number_composition = int(sheet.cell(row=i, column=39).value)
    student.essay = sheet.cell(row=i, column=55).value
    student.tokens = len(nltk.word_tokenize(student.essay))
    student.types = len(set(nltk.word_tokenize(student.essay)))
    #add the now populated Student oject to the list of students
    student_list.append(student)

# sort the student list based on placement score.  This is to used to calculate
# placement score percentile
student_list.sort(key=lambda x: x.placement_raw)
# an empty list to store all of the students' placement scores
score_list = list()
# add student placement scores to the score list
for student in student_list:
    score_list.append(student.placement_raw)
# calculate each student's placement score percentile, and save result to
# student's placement_percentile attribute.
for i in range(len(student_list)):
    score = score_list[i]
    percentile = stats.percentileofscore(score_list, score, 'rank')
    student_list[i].placement_percentile = percentile

# divide students into test and training sets.  Update each student's
# date_set attribute according to the data set into which the student is placed.
# One of every 15 students is places in test set.  All others placed in
# training set.
for i in range(len(student_list)):
    if i%15 == 0:
        student_list[i].data_set = 'test'
        student_test_list.append(student_list[i])
    else:
        student_list[i].data_set = 'training'
        student_train_list.append(student_list[i])

#save the data sets to file for futher use
pickle.dump(student_list, open('student_all.set', 'wb'))
pickle.dump(student_test_list, open('student_test.set', 'wb'))
pickle.dump(student_train_list, open('student_train.set', 'wb'))
